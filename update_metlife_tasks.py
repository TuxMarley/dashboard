import argparse
import datetime
import json
from pathlib import Path

import openpyxl


PROJECT_ROOT = Path(__file__).resolve().parent
DEFAULT_JSON_PATH = PROJECT_ROOT / "public" / "metlife_task_history.json"
SHEET_NAME = "Registro"
HEADER_ROW = 4
REQUIRED_HEADERS = {
    "Fecha",
    "Tipo de actividad",
    "Tarea / actividad",
    "Horas",
}


def parse_args():
    parser = argparse.ArgumentParser(
        description="Exporta el registro diario de MetLife a JSON para el dashboard."
    )
    parser.add_argument("--excel", type=Path, required=True, help="Ruta al libro de MetLife.")
    parser.add_argument(
        "--output",
        type=Path,
        default=DEFAULT_JSON_PATH,
        help="Archivo JSON de salida.",
    )
    return parser.parse_args()


def as_date(value):
    if isinstance(value, datetime.datetime):
        return value.date()
    if isinstance(value, datetime.date):
        return value
    if isinstance(value, str):
        return datetime.date.fromisoformat(value.strip())
    raise ValueError(f"Fecha no válida en la hoja {SHEET_NAME}: {value!r}")


def as_hours(value):
    if value is None or value == "":
        return 0
    hours = float(value)
    return int(hours) if hours.is_integer() else hours


def main():
    args = parse_args()

    if not args.excel.is_file():
        raise FileNotFoundError(f"No se encontró el libro: {args.excel}")

    workbook = openpyxl.load_workbook(args.excel, data_only=True, read_only=True)
    if SHEET_NAME not in workbook.sheetnames:
        raise KeyError(f"No existe la hoja requerida: {SHEET_NAME}")

    sheet = workbook[SHEET_NAME]
    headers = {
        str(cell.value).strip(): cell.column
        for cell in sheet[HEADER_ROW]
        if cell.value is not None
    }
    missing_headers = REQUIRED_HEADERS.difference(headers)
    if missing_headers:
        missing = ", ".join(sorted(missing_headers))
        raise KeyError(f"Faltan columnas requeridas en {SHEET_NAME}: {missing}")

    days = {}
    row_count = 0

    for row_index in range(HEADER_ROW + 1, sheet.max_row + 1):
        date_value = sheet.cell(row=row_index, column=headers["Fecha"]).value
        activity_type = sheet.cell(
            row=row_index, column=headers["Tipo de actividad"]
        ).value
        title = sheet.cell(row=row_index, column=headers["Tarea / actividad"]).value
        hours_value = sheet.cell(row=row_index, column=headers["Horas"]).value

        if all(value in (None, "") for value in (date_value, activity_type, title, hours_value)):
            continue
        if date_value in (None, "") or title in (None, ""):
            continue

        date = as_date(date_value)
        date_key = date.isoformat()
        tasks = days.setdefault(date_key, [])
        task_number = len(tasks) + 1
        tasks.append(
            {
                "id": f"{date_key}-{task_number}",
                "date": date_key,
                "type": str(activity_type or "Sin categoría").strip(),
                "title": str(title).strip(),
                "hours": as_hours(hours_value),
            }
        )
        row_count += 1

    ordered_days = {date: days[date] for date in sorted(days)}
    payload = {
        "version": 1,
        "updatedAt": datetime.datetime.now(datetime.timezone.utc).isoformat(),
        "days": ordered_days,
    }

    args.output.parent.mkdir(parents=True, exist_ok=True)
    with args.output.open("w", encoding="utf-8") as output:
        json.dump(payload, output, indent=2, ensure_ascii=False)
        output.write("\n")

    print(
        f"Exportados {row_count} registros de {len(ordered_days)} días a {args.output}"
    )


if __name__ == "__main__":
    main()
