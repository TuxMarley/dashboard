import openpyxl
import json
import datetime
import os

# Define file paths
excel_path = r"C:\Dev\Dashboard\documentos\avangrid\Native Mobile - CMP Regression Automation.xlsx"
json_output_path = r"C:\Dev\Dashboard\public\daily_tasks.json"

def main():
    if not os.path.exists(excel_path):
        print(f"Error: Excel file not found at {excel_path}")
        return

    # Load workbook
    wb = openpyxl.load_workbook(excel_path, data_only=False)
    
    # Get current date
    today = datetime.date.today()
    day_num_str = str(today.day)
    month_name = today.strftime("%B")  # e.g. "July"
    
    print(f"Scanning Excel for date: {today} ({month_name} {day_num_str})")
    
    sheets_to_scan = ['ManagePayments', 'MakePayment', 'Login', 'Autopay', 'Outages']
    today_tasks = []

    for name in sheets_to_scan:
        if name not in wb.sheetnames:
            continue
        ws = wb[name]
        
        # 1. Find the column for today
        today_col = None
        max_c = ws.max_column
        for c in range(1, max_c + 1):
            # Resolve month name for column c
            col_month = None
            for temp_c in range(c, 0, -1):
                m_val = ws.cell(row=1, column=temp_c).value
                if m_val is not None:
                    col_month = str(m_val).strip()
                    break
            
            col_day_num = ws.cell(row=3, column=c).value
            
            # Match current month and day number
            if col_month == month_name and (str(col_day_num) == day_num_str or col_day_num == today.day):
                today_col = c
                break
                
        if today_col is None:
            print(f"  Info: No column found for {month_name} {day_num_str} in sheet '{name}'")
            continue
            
        print(f"  Scanning sheet '{name}', found today's column at index {today_col}")
        
        # 2. Iterate rows starting from row 5
        for r in range(5, ws.max_row + 1):
            status_val = ws.cell(row=r, column=4).value
            if status_val == "QA Validation":
                cell = ws.cell(row=r, column=today_col)
                fill = cell.fill
                
                # Check for background fill (excluding default weekend pattern)
                has_fill = False
                color_val = None
                if fill and fill.fill_type and fill.fill_type != 'none':
                    # We check if fgColor is valid
                    if fill.fgColor:
                        color_val = fill.fgColor.value
                        # Color '00000000' or 0 is default / no fill in some styles, 
                        # but custom colors will have hex values like 'FFFFC000' (orange)
                        if color_val and str(color_val) not in ['00000000', '0', '000000']:
                            has_fill = True
                
                if has_fill:
                    tc_name = ws.cell(row=r, column=1).value
                    tc_key = ws.cell(row=r, column=2).value
                    assigned = ws.cell(row=r, column=3).value
                    
                    today_tasks.append({
                        "date": today.strftime("%Y-%m-%d"),
                        "sheet": name,
                        "key": tc_key,
                        "name": tc_name,
                        "assigned": assigned
                    })

    # Save results to JSON file
    os.makedirs(os.path.dirname(json_output_path), exist_ok=True)
    with open(json_output_path, "w", encoding="utf-8") as f:
        json.dump(today_tasks, f, indent=2, ensure_ascii=False)
        
    print(f"\nSuccessfully found {len(today_tasks)} tasks. Saved to {json_output_path}")

if __name__ == "__main__":
    main()
